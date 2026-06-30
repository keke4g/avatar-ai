import os
import numpy as np
from pathlib import Path
from pypdf import PdfReader
from docx import Document
import openpyxl

# Optional imports for SentenceTransformers
try:
    from sentence_transformers import SentenceTransformer
    EMBEDDING_MODEL = SentenceTransformer("all-MiniLM-L6-v2")
    HAS_TRANSFORMERS = True
    print("[INFO] Modelo SentenceTransformer 'all-MiniLM-L6-v2' cargado correctamente para RAG semántico.")
except ImportError:
    HAS_TRANSFORMERS = False
    print("[WARN] sentence-transformers no está instalado. Se utilizará una vectorización básica por frecuencias.")

class LocalVectorDB:
    def __init__(self):
        self.chunks = []      # List of strings
        self.embeddings = []  # List of numpy arrays

    def clear(self):
        self.chunks = []
        self.embeddings = []

    def get_embedding(self, text: str) -> np.ndarray:
        if HAS_TRANSFORMERS:
            # Generate premium semantic embedding
            vector = EMBEDDING_MODEL.encode(text, convert_to_numpy=True)
            return vector
        else:
            # Fallback simple letter-frequency frequency vector (matches original next.js behavior)
            vector = np.zeros(26, dtype=np.float32)
            clean_text = text.lower()
            for char in clean_text:
                code = ord(char)
                if 97 <= code <= 122:
                    vector[code - 97] += 1
            # Normalize vector to avoid divide-by-zero
            norm = np.linalg.norm(vector)
            if norm > 0:
                vector = vector / norm
            return vector

    def add_chunks(self, text_list: list[str]):
        for chunk in text_list:
            if not chunk.strip():
                continue
            emb = self.get_embedding(chunk)
            self.chunks.append(chunk)
            self.embeddings.append(emb)

    def search(self, query: str, top_k: int = 3) -> str:
        if not self.chunks:
            return ""

        query_emb = self.get_embedding(query)
        
        # Calculate cosine similarity using numpy
        scores = []
        for idx, emb in enumerate(self.embeddings):
            dot_product = np.dot(query_emb, emb)
            norm_q = np.linalg.norm(query_emb)
            norm_e = np.linalg.norm(emb)
            
            if norm_q > 0 and norm_e > 0:
                similarity = dot_product / (norm_q * norm_e)
            else:
                similarity = 0.0
                
            scores.append((similarity, self.chunks[idx]))

        # Sort by similarity descending
        scores.sort(key=lambda x: x[0], reverse=True)
        
        # Return joined top chunks
        relevant_chunks = [text for score, text in scores[:top_k]]
        return "\n\n".join(relevant_chunks)

# Initialize global Vector DB
vector_db = LocalVectorDB()

# Document Parsers
def extract_text_from_pdf(file_path: Path) -> str:
    text = ""
    try:
        reader = PdfReader(file_path)
        for page in reader.pages:
            t = page.extract_text()
            if t:
                text += t + "\n"
    except Exception as e:
        print(f"[ERROR] Error al leer PDF: {e}")
    return text

def extract_text_from_docx(file_path: Path) -> str:
    text = ""
    try:
        doc = Document(file_path)
        for para in doc.paragraphs:
            if para.text:
                text += para.text + "\n"
    except Exception as e:
        print(f"[ERROR] Error al leer DOCX: {e}")
    return text

def extract_text_from_xlsx(file_path: Path) -> str:
    text = ""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                row_str = " ".join([str(cell) for cell in row if cell is not None])
                if row_str.strip():
                    text += row_str + "\n"
    except Exception as e:
        print(f"[ERROR] Error al leer XLSX: {e}")
    return text

def chunk_text(text: str, chunk_size: int = 800, overlap: int = 100) -> list[str]:
    chunks = []
    if not text:
        return chunks
        
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]
        chunks.append(chunk)
        start += chunk_size - overlap
    return chunks

def process_and_index_file(file_path: str, filename: str) -> int:
    path = Path(file_path)
    if not path.exists():
        return 0

    suffix = path.suffix.lower()
    full_text = ""

    if suffix == ".pdf":
        full_text = extract_text_from_pdf(path)
    elif suffix == ".docx":
        full_text = extract_text_from_docx(path)
    elif suffix in [".xlsx", ".xls"]:
        full_text = extract_text_from_xlsx(path)
    elif suffix == ".txt":
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            full_text = f.read()
    else:
        print(f"[WARN] Formato no soportado en RAG: {suffix}")
        return 0

    chunks = chunk_text(full_text)
    if chunks:
        vector_db.clear()
        vector_db.add_chunks(chunks)
        print(f"[INFO] Indexados {len(chunks)} chunks del archivo: {filename}")
        return len(chunks)
    
    return 0
